
import pandas
import openpyxl
import os


from bs4 import BeautifulSoup
import requests


def soup(me):
    """
    :param me: element
    :return:
    """
    me = str(me)
    btf = BeautifulSoup(me, 'html.parser')
    return btf


class Pandas:
    my_file = 'ana.xlsx'


inst = Pandas()

mcp = openpyxl.load_workbook(inst.my_file)
# mycopy.create_sheet('Socios')

wks = mcp.worksheets
ws1 = wks[0]
# ws2 = wks[1]

# input(f1st['A1'].value)
"""
for e in sheet1['A1:E20']:
    for i in range(10):
        print(f'A{i}')
"""
__dict = {}
# campos = []
cnpjs = ["47.960.950/0001-21",
         "15.436.940/0001-03",
         "00.776.574/0006-60",
         "33.041.260/0652-90",
         "47.960.950/0001-21",
         "45.543.915/0846-95",
         "33.041.260/0652-90"]
# "".join([a for a in c if a.isnumeric()])
newcnpjs = []
for c in cnpjs:
    newc = ""
    for part in c:
        if part.isnumeric():
            newc += part
    newcnpjs.append(newc)
cnpjs.clear()
cnpjs = newcnpjs.copy()

for cont, cnpj in enumerate(cnpjs[:]):

    try:
        import re
        # CADASTRA CNAE / DESCRIÇÃO
        now_link = f'http://cnpj.info/{cnpj}'
        # driver.get(now_link)
        # cnae = driver.find_element_by_tag_name("u").text

        req = requests.get(now_link).text
        thesoup = soup(req)
        """
        # o select gera uma lista, aceitando, indexação ou for... etc
        cnae_total = thesoup.select("u")[0].text
        # cnae_total significa só "-"
        # RSPLIT
        cnae, descricao = cnae_total.rsplit(cnae_total[11], 1)
        cnae = cnae.strip()
        descricao = descricao.strip()
        print(cnae, descricao, '------>', cnpj)
        ws1[f"J{str(cont+2)}"].value = cnae
        ws1[f"K{str(cont+2)}"].value = descricao
        """
        # endereço
        # document.querySelector("body > div.container > div > div > div.col.c9-2 > p:nth-child(14)")
        endereco = thesoup.find('h3', text=re.compile(
            'Endereço'),)  # attrs={'class': 'pos'})
        try:
            items = endereco.next_siblings
        except AttributeError:
            print('teste atr')
            continue
        print('teste')
        _dados_header = ['Logadouro', 'Complemento',
                         'Bairro', 'Cidade e Estado', 'CEP']
        dados = []
        for e, txt in enumerate(items):
            txt = str(txt).strip()
            if e == 9:
                break
            if 'br' not in txt:
                print(txt)
                dados.append(txt) if txt != '' else None

        if '<h3>Contatos</h3>' == dados[-1]:
            dados.insert(1, '-')
        # dados passou de list p/ dict abaixo
        dados = {k: v for k, v in zip(_dados_header, dados)}
        try:
            dados["Cidade"], dados["Estado"] = dados['Cidade e Estado'].split(
                '-')
        except ValueError:
            input(
                f"NÃO DEU CERTO, LINHA dados['Cidade e Estado']: {dados['Cidade e Estado']}")

        del dados["Cidade e Estado"]
        del _dados_header[-2]
        _dados_header.append("Cidade")
        _dados_header.append("Estado")
        _dados_header.insert(0, "CNPJ")

        dados["CNPJ"] = cnpj
        for row in ws1.iter_rows(cont+2, cont+2, 12, 17):
            cont_cell = 0
            for cell in row:
                # if cont_cell == 3:
                try:
                    cell.value = dados[_dados_header[cont_cell]]
                except KeyError:
                    input(f'inputing: {dados}')
                cont_cell += 1

            # el = thesoup.find(f"body > div.container > div > div > div.col.c9-2 > p:nth-child({i})")
        # print(el[0].text)
    except IndexError:
        raise IndexError


mcp.save(inst.my_file)
print('finish')
